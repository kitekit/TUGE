#coding:utf-8
'''
    此程序用于设备流量统计分析
    使用索引能大幅提升原始流量表向二维统计表的数据迁移
    使用sys.path[0]自动适配当前文件夹，使程序更为通用
    自动测算统计表中的设备数和统计日期个数，减少自定义常量，使程序更为通用
    
    输入值包括：
    1、flow.xlsx  BSS导出的设备天流量表，核心字段是设备号，日期和流量值
    2、sim.xlsx   BSS导出的当前设备与运营商卡绑定关系表，核心字段是设备号，运营商卡名称，sim卡号，SIM卡流量
    3、statistics.xlsx  统计分析表，预先明确分析的设备号和分析时间段

    输出值均写入statistics.xlsx
    1、流量值匹配
    2、流量统计值
    3、sim卡绑定匹配关系分析
    4、建议互换卡的设备号对  
'''

import sys
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
#使用sys.path[0]获取当前路径
flow=load_workbook(sys.path[0]+r'\flow.xlsx') 
sim= load_workbook(sys.path[0]+r'\sim.xlsx')                    
statistics=load_workbook(sys.path[0]+r'\statistics.xlsx')
flow_sheet=flow.active
sim_sheet=sim.active
statistics_sheet=statistics.active
lst_statistics_col=[]
lst_statistics_row=[]
#为设备号做索引
for statistics_sheet_row in range(2,65535):                           
    if statistics_sheet.cell(statistics_sheet_row,1).value==None:
        break
    lst_statistics_row.append(statistics_sheet.cell(statistics_sheet_row,1).value)
#为日期做索引
for statistics_sheet_col in range(2,1000):
    if statistics_sheet.cell(1,statistics_sheet_col).value==None:
        break
    lst_statistics_col.append(statistics_sheet.cell(1,statistics_sheet_col).value)
for flow_sheet_row in range(2,65535):
    try:
#执行到流量表为空的一行，自动退出循环
        if flow_sheet.cell(flow_sheet_row,1).value==None:            
           break
#根据设备号和日期获取统计表中的行列号，直接赋值
        x=lst_statistics_row.index(flow_sheet.cell(flow_sheet_row,1).value)
        y=lst_statistics_col.index(flow_sheet.cell(flow_sheet_row,4).value)
        statistics_sheet.cell(x+2,y+2).value=flow_sheet.cell(flow_sheet_row,3).value
#由于流量表中部分记录不在索引表中，相关操作可能返回异常，需增加异常处理
    except:
        pass
#将sim绑卡表中信息导入statistics统计表 流量列后第九列
for sim_row in range(2,65535):
    if sim_sheet.cell(sim_row,7).value==None:
        break
    elif sim_sheet.cell(sim_row,7).value in lst_statistics_row:
        x=lst_statistics_row.index(sim_sheet.cell(sim_row,7).value)+2
        y=len(lst_statistics_col)+1+9
        statistics_sheet.cell(x,y).value=sim_sheet.cell(sim_row,5).value
    else:
        pass
#基于第一个非空记录且大于10M的记录,确定设备是从第几天开始启用,记录在流量列后第八列
for statistics_sheet_row in range(2,65535): 
    if statistics_sheet.cell(statistics_sheet_row,1).value==None:
        break
    for statistics_sheet_col in range(2,1000):
        if statistics_sheet.cell(1,statistics_sheet_col).value==None:
            break
        if statistics_sheet.cell(statistics_sheet_row,statistics_sheet_col).value!=None:
            if statistics_sheet.cell(statistics_sheet_row,statistics_sheet_col).value>10:
                y=len(lst_statistics_col)+1+8
                statistics_sheet.cell(statistics_sheet_row,y).value=statistics_sheet_col-1
                break
            else:
                pass
        else:
            pass

#将内存表进行存储
statistics.save(sys.path[0]+r'\statistics.xlsx')
